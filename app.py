from flask import Flask, request, abort
from linebot import LineBotApi, WebhookHandler
from linebot.models import *
from linebot.exceptions import InvalidSignatureError
from openpyxl import Workbook, load_workbook
import os
import datetime

app = Flask(__name__)

CHANNEL_ACCESS_TOKEN = "G+WVfBYA61JkTqXM+r1/Y3HlksAxH2sq8lm/Zm2ifIr6agVcr339zAbjUfmZSp5tkyB97l0pG681K7hGtUnbmXnV2CyG8O7t/dTDhoA9CJI94aftXczCEfFwmWTTEu3tdLWNDWiHH7lzrW4uPX0UOgdB04t89/1O/w1cDnyilFU="
CHANNEL_SECRET = "ab06793a22d874528049156bf4de7dc8"

line_bot_api = LineBotApi(CHANNEL_ACCESS_TOKEN)
handler = WebhookHandler(CHANNEL_SECRET)

ADMIN_USER_ID = "U6b88cd42924581ac141db052cc09eb08"

REGIONS = ["福岡", "熊本", "大分", "佐賀", "長崎"]
USER_REGION = {}

FILE_NAME = "health_log.xlsx"

# =========================
# Excel保存
# =========================
def save_log(user_id, region, message):
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.append(["日時", "ユーザーID", "地域", "内容"])
        wb.save(FILE_NAME)

    wb = load_workbook(FILE_NAME)
    ws = wb.active
    ws.append([now, user_id, region, message])
    wb.save(FILE_NAME)

# =========================
# アラート
# =========================
def send_alert(user_id, region, message):
    alert = f"🚨【緊急アラート】\nユーザー: {user_id}\n地域: {region}\n内容: {message}"
    line_bot_api.push_message(ADMIN_USER_ID, TextSendMessage(text=alert))

@app.route("/")
def index():
    return "健康チェックBOT起動中"

@app.route("/callback", methods=["POST"])
def callback():
    signature = request.headers["X-Line-Signature"]
    body = request.get_data(as_text=True)
    handler.handle(body, signature)
    return "OK"

@handler.add(MessageEvent, message=TextMessage)
def handle_message(event):
    user_id = event.source.user_id
    text = event.message.text.strip()

    region = USER_REGION.get(user_id, "未登録")

    # ログ保存
    save_log(user_id, region, text)

    # アラート
    if text in ["かなり悪いです", "メンタル_いいえ", "ケガあり"]:
        send_alert(user_id, region, text)

    # =========================
    # 緊急開始
    # =========================
    if text == "緊急開始":
        buttons = QuickReply(items=[
            QuickReplyButton(action=MessageAction(label="ケガあり", text="ケガあり")),
            QuickReplyButton(action=MessageAction(label="危険", text="危険")),
            QuickReplyButton(action=MessageAction(label="安全", text="安全")),
        ])
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text="現在の状況を選んでください", quick_reply=buttons)
        )
        return

    # =========================
    # ★ケガあり（GPS説明強化）
    # =========================
    if text == "ケガあり":
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(
                text="それは大変です。救急車を呼べますか？誰か近くにいますか？\n\n可能であればGPSを送ってください。\n\n【送り方】\n①入力欄の文字をすべて消してください\n②左の「＋」ボタンを押してください\n③「位置情報」を選択してください\n④表示された地図でそのまま送信を押してください"
            )
        )
        return

    # =========================
    # ★危険（GPS説明強化）
    # =========================
    if text == "危険":
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(
                text="危険な状況です。安全な場所に避難してください。\n周りの人と一緒に行動してください\n\n可能であればGPSを送ってください。\n\n【送り方】\n①入力欄の文字をすべて消してください\n②左の「＋」ボタンを押してください\n③「位置情報」を選択してください\n④表示された地図でそのまま送信を押してください"
            )
        )
        return

    if text == "安全":
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text="安全が確認できて安心しました。\n引き続き注意して行動してください")
        )
        return

    # =========================
    # 通常開始
    # =========================
    if text == "開始":
        buttons = QuickReply(items=[
            QuickReplyButton(action=MessageAction(label=r, text=r)) for r in REGIONS
        ])
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text="地域を選択してください", quick_reply=buttons)
        )
        return

    # 地域登録
    if text in REGIONS:
        USER_REGION[user_id] = text
        region = text

        buttons = QuickReply(items=[
            QuickReplyButton(action=MessageAction(label="元気です", text="元気です")),
            QuickReplyButton(action=MessageAction(label="少し元気です", text="少し元気です")),
            QuickReplyButton(action=MessageAction(label="かなり悪いです", text="かなり悪いです")),
        ])

        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text=f"{text}で登録しました。\n\n🌟 元気ですか？ 🌟", quick_reply=buttons)
        )
        return

    # =========================
    # 健康チェック
    # =========================
    if text == "元気です":
        line_bot_api.reply_message(event.reply_token,
            TextSendMessage(text="よかったです！今日も一日頑張りましょう"))
        return

    if text == "少し元気です":
        buttons = QuickReply(items=[
            QuickReplyButton(action=MessageAction(label="はい", text="少し元気_はい")),
            QuickReplyButton(action=MessageAction(label="いいえ", text="少し元気_いいえ")),
        ])
        line_bot_api.reply_message(event.reply_token,
            TextSendMessage(text="大丈夫ですか？", quick_reply=buttons))
        return

    if text == "少し元気_はい":
        line_bot_api.reply_message(event.reply_token,
            TextSendMessage(text="無理せず様子を見てください"))
        return

    if text == "少し元気_いいえ":
        line_bot_api.reply_message(event.reply_token,
            TextSendMessage(text="無理せず休んでください"))
        return

    if text == "かなり悪いです":
        buttons = QuickReply(items=[
            QuickReplyButton(action=MessageAction(label="体調", text="体調")),
            QuickReplyButton(action=MessageAction(label="メンタル", text="メンタル")),
        ])
        line_bot_api.reply_message(event.reply_token,
            TextSendMessage(text="原因は何ですか？", quick_reply=buttons))
        return

    if text == "体調":
        line_bot_api.reply_message(event.reply_token,
            TextSendMessage(text="無理せずに病院に行かれてください"))
        return

    if text == "メンタル":
        buttons = QuickReply(items=[
            QuickReplyButton(action=MessageAction(label="はい", text="メンタル_はい")),
            QuickReplyButton(action=MessageAction(label="いいえ", text="メンタル_いいえ")),
        ])
        line_bot_api.reply_message(event.reply_token,
            TextSendMessage(text="大丈夫ですか？", quick_reply=buttons))
        return

    if text == "メンタル_はい":
        line_bot_api.reply_message(event.reply_token,
            TextSendMessage(text="無理せず様子を見てください"))
        return

    if text == "メンタル_いいえ":
        buttons = QuickReply(items=[
            QuickReplyButton(action=MessageAction(label="個人", text="個人")),
            QuickReplyButton(action=MessageAction(label="仕事", text="仕事")),
        ])
        line_bot_api.reply_message(event.reply_token,
            TextSendMessage(text="原因はどちらですか？", quick_reply=buttons))
        return

    if text == "個人":
        line_bot_api.reply_message(event.reply_token,
            TextSendMessage(text="一人で悩まず誰かに相談してください"))
        return

    if text == "仕事":
        line_bot_api.reply_message(event.reply_token,
            TextSendMessage(text="委員長が相談に乗ります。まず電話してください"))
        return